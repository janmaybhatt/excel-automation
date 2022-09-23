from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import random


if __name__ == '__main__':
    """wb = Workbook()
    sheet = wb.active
    email = sheet.cell(row=1, column=1)
    email.value = 'Email'
    username = sheet.cell(row=1, column=2)
    username.value = 'Username'
    password = sheet.cell(row=1, column=3)
    password.value = 'Password'
    wb.save('test.xlsx')"""
    wb = load_workbook('test.xlsx')
    sheet = wb.active
    green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    blue = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    data = []
    for i in range(10):
        col_value = random.choice(['Started', 'In Progress', 'Completed'])
        data.append([
            ['jrbhatt18p10@gmail.com', 'janmaybhatt', col_value]
        ])
    row = sheet.max_row
    for i in range(row, row + len(data)):
        for j in range(1, 4):
            sheet.cell(row=i, column=j).value = data[i - row][0][j - 1]
            if sheet.cell(row=i, column=j).value == 'Started':
                sheet.cell(row=i, column=j).fill = green
            elif sheet.cell(row=i, column=j).value == 'In Progress':
                sheet.cell(row=i, column=j).fill = blue
            elif sheet.cell(row=i, column=j).value == 'Completed':
                sheet.cell(row=i, column=j).fill = red
    wb.save('test.xlsx')
